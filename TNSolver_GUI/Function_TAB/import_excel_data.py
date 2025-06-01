"""
    This class defines the Frame for importing data prepared via Microsoft Excel

    Luca Lombardi
    29 May 2025: First Draft

    next steps:

"""
from tkinter import Tk, Frame, filedialog, messagebox, LabelFrame, Label, Button
from tkinter.ttk import Treeview, Scrollbar, Combobox
from TNSolver_GUI.Thermal_Network_TAB.gUtility import is_float
from openpyxl import load_workbook


class ExcelImporterApp(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.parent = parent  # Store the parent reference
        self.workbook = None
        self.sheet = None
        self.header_row = []
        self.all_data = []  # Stores all data as a list of lists (values only)
        self.data_dictionary = {}  # Stores data from selected columns

        #  main frame
        self._import_excel_frame = LabelFrame(self, text="Microsoft Excel data import", padx=10, pady=10)
        self._import_excel_frame.pack(side='left', padx=10, pady=10, expand=True, fill='x', anchor='n')

        # --- Frame for File Selection ---
        self.file_frame = LabelFrame(self._import_excel_frame, text="Select Excel File")
        self.file_frame.pack(pady=10, padx=10, fill="x")

        self.file_path_label = Label(self.file_frame, text="No file selected")
        self.file_path_label.pack(side="left", padx=5, pady=5)

        self.browse_button = Button(self.file_frame, text="Browse", command=self.browse_excel_file)
        self.browse_button.pack(side="right", padx=5, pady=5)

        # --- Frame for Data Display ---
        self.data_frame = LabelFrame(self._import_excel_frame, text="Cleaned Sheet Data")
        self.data_frame.pack(pady=10, padx=10, fill="both", expand=True)

        self.tree = Treeview(self.data_frame, show="headings")
        self.tree.pack(side="left", fill="both", expand=True)

        # Scrollbars for the Treeview
        self.vsb = Scrollbar(self.data_frame, orient="vertical", command=self.tree.yview)
        self.vsb.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=self.vsb.set)

        self.hsb = Scrollbar(self.data_frame, orient="horizontal", command=self.tree.xview)
        self.hsb.pack(side='bottom', fill='x')
        self.tree.configure(xscrollcommand=self.hsb.set)

        # --- Frame for Column Selection ---
        self.column_selection_frame = LabelFrame(self._import_excel_frame, text="Select Two Columns")
        self.column_selection_frame.pack(pady=10, padx=10, fill="x")

        self.column1_label = Label(self.column_selection_frame, text="Coordinate:")
        self.column1_label.pack(side="left", padx=5, pady=5)
        self.column1_combobox = Combobox(self.column_selection_frame, state="readonly")
        self.column1_combobox.pack(side="left", padx=5, pady=5, expand=True, fill="x")

        self.column2_label = Label(self.column_selection_frame, text="Data value:")
        self.column2_label.pack(side="left", padx=5, pady=5)
        self.column2_combobox = Combobox(self.column_selection_frame, state="readonly")
        self.column2_combobox.pack(side="left", padx=5, pady=5, expand=True, fill="x")

        self.import_columns_button = Button(self.column_selection_frame, text="Import Selected Columns",
                                            command=self.import_selected_columns)
        self.import_columns_button.pack(side="right", padx=5, pady=5)

    def browse_excel_file(self):
        # Set the parent window as transient for the file dialog
        self.parent.transient(self.parent.master)  # This is crucial if self.parent is a Toplevel
        self.parent.grab_set()  # Grab focus of the parent window

        file_path = filedialog.askopenfilename(
            parent=self.parent,  # Explicitly set the parent for the file dialog
            filetypes=[("Excel files", "*.xlsx")])

        if file_path:
            self.file_path_label.config(text=file_path)
            self.load_excel_data(file_path)

        self.parent.grab_release()  # Release the grab after the dialog closes
        # Bring the parent window to the front again after the dialog closes
        self.parent.lift()

    def load_excel_data(self, file_path):
        try:
            self.workbook = load_workbook(filename=file_path, data_only=True)
            self.sheet = self.workbook.active

            self.all_data = []
            self.header_row = []

            # --- Find the first non-empty row and column ---
            min_row_data = 0
            for row_idx, row in enumerate(self.sheet.iter_rows(values_only=True)):
                if any(cell_value is not None for cell_value in row):
                    min_row_data = row_idx + 1  # +1 because iter_rows is 0-indexed, but sheet is 1-indexed
                    break

            if min_row_data == 0:  # No data found at all
                messagebox.showwarning("Warning", "The selected Excel sheet is empty or contains only empty rows.")
                self.clear_all_data()
                return

            # Find the first non-empty column based on the "data region"
            min_col_data = self.sheet.max_column + 1  # Initialize with a value beyond max
            for col_idx in range(1, self.sheet.max_column + 1):  # Iterate through all potential columns
                has_data_in_col = False
                for row_idx in range(min_row_data, self.sheet.max_row + 1):
                    cell_value = self.sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        has_data_in_col = True
                        break
                if has_data_in_col:
                    min_col_data = col_idx
                    break

            if min_col_data > self.sheet.max_column:  # No data found in any column from the first non-empty row
                messagebox.showwarning("Warning",
                                       "The selected Excel sheet does not contains data.")
                self.clear_all_data()
                return

            # Now, read the data from the identified data region
            # max_column and max_row still give the overall max, which is fine here
            # since min_col_data and min_row_data narrow down the starting point.
            num_columns = self.sheet.max_column
            num_rows = self.sheet.max_row  # Use actual max_row, not just first data row

            # Read all rows from the identified starting row and column
            all_rows_raw = list(self.sheet.iter_rows(min_row=min_row_data,
                                                     max_row=num_rows,
                                                     min_col=min_col_data,
                                                     max_col=num_columns,
                                                     values_only=True))

            if not all_rows_raw:
                messagebox.showwarning("Warning", "No data found after cleaning initial empty rows/columns.")
                self.clear_all_data()
                return

            # Assume the first row of the cleaned data is the header
            self.header_row = [str(cell_value) if cell_value is not None else "" for cell_value in all_rows_raw[0]]

            # Store the rest of the data
            self.all_data = []
            for row_data_tuple in all_rows_raw[1:]:  # Skip the header row
                self.all_data.append(
                    [str(cell_value) if cell_value is not None else "" for cell_value in row_data_tuple])

            self.display_data_in_treeview()
            self.populate_column_comboboxes()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")
            self.clear_all_data()

    def clear_all_data(self):
        self.workbook = None
        self.sheet = None
        self.header_row = []
        self.all_data = []
        self.clear_treeview()
        self.clear_column_comboboxes()

    def display_data_in_treeview(self):
        self.clear_treeview()
        if self.header_row:
            self.tree["columns"] = self.header_row
            for col in self.header_row:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100, anchor="w")

            for row_data in self.all_data:
                self.tree.insert("", "end", values=row_data)

    def clear_treeview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.tree["columns"] = ()

    def populate_column_comboboxes(self):
        if self.header_row:
            self.column1_combobox["values"] = self.header_row
            self.column2_combobox["values"] = self.header_row
            self.column1_combobox.set("")
            self.column2_combobox.set("")
        else:
            self.clear_column_comboboxes()

    def clear_column_comboboxes(self):
        self.column1_combobox["values"] = []
        self.column2_combobox["values"] = []
        self.column1_combobox.set("")
        self.column2_combobox.set("")

    def import_selected_columns(self):
        col1_name = self.column1_combobox.get()
        col2_name = self.column2_combobox.get()

        if not self.all_data:
            messagebox.showwarning("Warning", "No data loaded. Please import an Excel file first.")
            return

        if not (col1_name and col2_name):
            messagebox.showwarning("Warning", "Please select two columns first.")
            return

        if col1_name == col2_name:
            messagebox.showwarning("Warning", "Please select two different columns.")
            return

        col1_index = self.header_row.index(col1_name)
        col2_index = self.header_row.index(col2_name)

        self.data_dictionary = {}
        err_number = 0

        for row in self.all_data:
            value1 = row[col1_index] if len(row) > col1_index else 'empty'
            value2 = row[col2_index] if len(row) > col2_index else 'empty'

            if is_float(value1) and is_float(value2):
                self.data_dictionary[float(value1)] = float(value2)
            else:
                err_number += 1

        if bool(self.data_dictionary): # check for an empty dictionary
            message = ('Data from {} and {} imported successfully!\n {} empty or text values skipped'
                       .format(col1_name, col2_name, err_number))
            messagebox.showinfo('Success!', message)
            return self.data_dictionary
        else:
            messagebox.showinfo('Error!', 'No numerical values found, check the file!')


if __name__ == "__main__":
    win = Tk()
    win.title('Test of importing Microsoft Excel file class')
    app = ExcelImporterApp(win)
    app.pack()
    win.mainloop()
